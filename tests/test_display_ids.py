"""
Tests for the two identifiers a task has, and which of them is on screen.

WHY THIS MODULE EXISTS:
======================
A task now carries two numbers that used to be one.

Task.id is the identity. Dependencies, parents, the clipboard, the tree's row
ids and every entry in the undo history are keyed on it, and it never changes
because a row moved. It is also never shown: it is a key, and a key in the
column where every other row shows its position would be read as a position.

The display id is where the row currently sits, counted from one down the
list. It is worked out rather than stored, which is the whole design - a
stored number would have to be rewritten on every reorder, insert, delete and
indent, and each of those is already recorded in the undo history against
Task.id. Renumbering a stored field after the change would leave that history
pointing at numbers naming nothing.

So the tests below are in two halves: that the number cascades exactly as the
specification's table says, and that the identity underneath it does not move
while it happens.

Nothing here needs a display.
"""

import unittest
from datetime import datetime, timedelta

from gantt_app.models import Project, Task

BASE = datetime(2026, 8, 25)


class DisplayIdTestCase(unittest.TestCase):
    """A plan whose tasks are named for what they are."""

    def plan(self, *rows) -> Project:
        """A project from (id, parent) pairs, in the order given."""
        project = Project(name="Plan")
        for task_id, parent in rows:
            project.add_task(Task(id=task_id, name=task_id, task_type="Task",
                                  parent_task_id=parent, start_date=BASE,
                                  end_date=BASE + timedelta(days=1)))
        return project

    def shown(self, project) -> list:
        """The tasks in display order, by identity."""
        return [task.id for task in project.display_order()]


class TestTheNumbersAreContiguous(DisplayIdTestCase):
    """One to N, with no gaps, whatever the plan has been through."""

    def test_a_flat_plan_counts_from_one(self):
        """The top of the list is 1, not 0."""
        project = self.plan(('a', None), ('b', None), ('c', None))

        self.assertEqual(project.display_ids(), {'a': 1, 'b': 2, 'c': 3})

    def test_children_are_counted_where_they_are_shown(self):
        """A sub-task's number follows its parent's, not the list order."""
        project = self.plan(('a', None), ('b', None), ('child', 'a'))

        self.assertEqual(project.display_ids(),
                         {'a': 1, 'child': 2, 'b': 3})

    def test_the_padded_form_is_what_the_column_shows(self):
        """Zero-padded, the way the list has always written a task number."""
        project = self.plan(('a', None), ('b', None))

        self.assertEqual(project.display_id('b'), '002')

    def test_a_task_that_is_not_in_the_plan_has_no_number(self):
        """Rather than an exception, or a misleading zero."""
        self.assertEqual(self.plan(('a', None)).display_id('gone'), '')


class TestTheSpecificationsTable(DisplayIdTestCase):
    """The three behaviours the specification sets out row by row."""

    def test_inserting_between_pushes_the_rest_down(self):
        """
        Task 1, Task 2 becomes Task 1, the new Task 2, Task 3.

        The new row is appended to the plan and shown under its parent, so
        this is also the case a numbering that counted list order gets
        wrong.
        """
        project = self.plan(('first', None), ('second', None))
        project.add_task(Task(id='inserted', name='New', task_type='Task',
                              parent_task_id='first', start_date=BASE,
                              end_date=BASE + timedelta(days=1)))

        self.assertEqual(project.display_ids(),
                         {'first': 1, 'inserted': 2, 'second': 3})

    def test_moving_a_row_up_swaps_the_two_numbers(self):
        """Task 3 dragged above Task 2 becomes Task 2, and it becomes 3."""
        project = self.plan(('one', None), ('two', None), ('three', None))

        project.move_task_before('three', 'two')

        self.assertEqual(self.shown(project), ['one', 'three', 'two'])
        self.assertEqual(project.display_ids(),
                         {'one': 1, 'three': 2, 'two': 3})

    def test_deleting_leaves_no_gap(self):
        """The rows below collapse up rather than the numbers skipping one."""
        project = self.plan(('one', None), ('two', None), ('three', None))

        project.remove_task('two')

        self.assertEqual(project.display_ids(), {'one': 1, 'three': 2})

    def test_indenting_renumbers_what_moved_past_it(self):
        """
        A hierarchy change alters the display order, so it alters the
        numbers - which is the fourth trigger the specification names.
        """
        project = self.plan(('a', None), ('b', None), ('c', None))

        project.indent_tasks(['c'])

        self.assertEqual(project.display_ids()['c'], 3)
        self.assertEqual(project.get_task_by_id('c').parent_task_id, 'b')


class TestTheIdentityDoesNotMove(DisplayIdTestCase):
    """What the numbering must not disturb."""

    def test_a_reorder_leaves_every_identity_alone(self):
        """
        Which is what lets the undo history survive it.

        Every entry in that history is keyed on Task.id. A renumbering that
        rewrote those keys would leave undo restoring an order of rows that
        had stopped existing.
        """
        project = self.plan(('one', None), ('two', None), ('three', None))
        before = {task.id for task in project.tasks}

        project.move_task_before('three', 'two')

        self.assertEqual({task.id for task in project.tasks}, before)

    def test_a_dependency_still_points_at_the_same_task(self):
        """
        The link is on the identity, so it survives the row moving.

        The *number* it is shown as changes, which is the specification's
        "visual predecessor references adjust dynamically".
        """
        project = self.plan(('one', None), ('two', None), ('three', None))
        project.get_task_by_id('three').add_dependency('one')

        project.move_task_before('three', 'two')

        link = project.get_task_by_id('three').dependencies[0]
        self.assertEqual(link.task_id, 'one')
        self.assertEqual(project.display_ids()[link.task_id], 1)

    def test_a_parent_reference_survives_a_renumber(self):
        """Hierarchy is held by identity too."""
        project = self.plan(('parent', None), ('child', 'parent'),
                            ('other', None))

        project.move_task_before('other', 'parent')

        self.assertEqual(project.get_task_by_id('child').parent_task_id,
                         'parent')

    def test_nothing_is_stored_to_go_stale(self):
        """
        There is no display id on the task to be written or forgotten.

        A field would need renumbering on four separate triggers; a number
        worked out from where the row sits is right the moment it moves.
        """
        task = self.plan(('a', None)).tasks[0]

        self.assertFalse(hasattr(task, 'display_id'))

    def test_the_saved_file_carries_the_identity(self):
        """
        And not the number, which would be a copy of a fact about order.

        A stored number in a file is one more thing that can disagree with
        the plan it came in.
        """
        data = self.plan(('a', None)).tasks[0].to_dict()

        self.assertEqual(data['id'], 'a')
        self.assertNotIn('display_id', data)


class TestWhatIsOnScreen(DisplayIdTestCase):
    """The identity is a key, and keys are not shown to readers."""

    def test_the_search_finds_a_row_by_the_number_shown(self):
        """Both the bare number and the padded form the column uses."""
        from gantt_app.views.searchbox import task_haystack

        project = self.plan(('first', None), ('second', None))
        text = task_haystack(project.get_task_by_id('second'), project)

        self.assertIn('2', text.split())
        self.assertIn('002', text.split())

    def test_the_search_does_not_find_it_by_identity(self):
        """
        Nobody can see the identity, so nobody can type it.

        Matching on it would find rows by a number that is nowhere on
        screen, and the number that is on screen would find the wrong row.
        """
        from gantt_app.views.searchbox import task_haystack

        project = self.plan(('ZZTOP', None))
        # Named differently from its identity, or the name would put the
        # identity into the haystack and the test would prove nothing
        project.tasks[0].name = "Kickoff"
        text = task_haystack(project.tasks[0], project)

        self.assertIn('kickoff', text)
        self.assertNotIn('zztop', text)



class TestTheExportsCarryTheNumber(DisplayIdTestCase):
    """
    An exported file names a task by the number the reader can see.

    Not by Task.id: that is a key nobody is shown, so a file naming a task
    by it names it by something the reader cannot look up in the plan they
    exported. Somebody reading a GanttProject file against the task list is
    looking for the row the list calls 4.
    """

    def exportable(self) -> Project:
        """A plan whose identities are nothing like its numbers."""
        project = Project(name="Demo")
        for task_id, parent in (('zz-1', None), ('aa-2', 'zz-1'),
                                ('mm-3', None)):
            project.add_task(Task(id=task_id, name=task_id.upper(),
                                  task_type="Task", parent_task_id=parent,
                                  start_date=BASE,
                                  end_date=BASE + timedelta(days=2)))
        project.get_task_by_id('mm-3').add_dependency('zz-1')
        return project

    def test_the_shared_walk_numbers_by_the_display_id(self):
        """
        Which is what both XML exporters write, so neither can drift.

        The walk used to keep a count of its own. It agreed with the list
        by construction and there were two definitions of the order, which
        is one more than can be kept in step.
        """
        from gantt_app.utils.plan_export import numbering, outline

        project = self.exportable()

        self.assertEqual(numbering(outline(project)), project.display_ids())

    def test_the_ganttproject_file_uses_it(self):
        """Both for the task ids and for the links between them."""
        import xml.etree.ElementTree as ET

        from gantt_app.utils.gan_exporter import generate_gan_content

        project = self.exportable()
        root = ET.fromstring(generate_gan_content(project))
        shown = project.display_ids()

        self.assertEqual([element.get('id') for element in root.iter('task')],
                         [str(shown[task.id])
                          for task in project.display_order()])

    def test_the_microsoft_file_uses_it(self):
        """UID and ID alike, which is what its links point at."""
        import xml.etree.ElementTree as ET

        from gantt_app.utils.msproject_exporter import (
            MSPDI_NAMESPACE, generate_msproject_content,
        )

        project = self.exportable()
        namespace = {'ms': MSPDI_NAMESPACE}
        root = ET.fromstring(generate_msproject_content(project))
        shown = project.display_ids()

        found = [task.find('ms:UID', namespace).text
                 for task in root.findall('ms:Tasks/ms:Task', namespace)]
        self.assertEqual(found, [str(shown[task.id])
                                 for task in project.display_order()])

    def test_no_identity_reaches_either_file(self):
        """The check that catches one leaking through a field nobody thought of."""
        from gantt_app.utils.gan_exporter import generate_gan_content
        from gantt_app.utils.msproject_exporter import generate_msproject_content

        project = self.exportable()
        written = generate_gan_content(project) + generate_msproject_content(project)

        for task in project.tasks:
            self.assertNotIn(task.id, written, task.id)

    def test_the_spreadsheet_uses_it_too(self):
        """
        Including the column its predecessor references point at.

        The sheet holds one row per piece of work and shows the phases as a
        column beside them, so these numbers have gaps where the phases sit.
        That is the right way round: a sheet read against the plan has to
        call a task what the plan calls it.
        """
        from gantt_app.utils.xlsx_exporter import (
            OPENPYXL_AVAILABLE, generate_xlsx_bytes,
        )

        if not OPENPYXL_AVAILABLE:
            self.skipTest("openpyxl is not installed")

        import io

        import openpyxl

        project = self.exportable()
        sheet = openpyxl.load_workbook(
            io.BytesIO(generate_xlsx_bytes(project))).worksheets[0]

        shown = project.display_ids()
        written = {sheet.cell(row=row, column=1).value
                   for row in range(6, sheet.max_row + 1)
                   if sheet.cell(row=row, column=1).value is not None}

        self.assertTrue(written, "the sheet wrote no ID column")
        self.assertTrue(written <= set(shown.values()),
                        f"{written} are not numbers the plan shows")

if __name__ == '__main__':
    unittest.main()
