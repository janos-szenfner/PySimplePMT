"""
Tests for the XLSX importer.
"""

import unittest
import tempfile
import os
from datetime import datetime

from gantt_app.utils.xlsx_importer import (
    XLSXImporter, import_xlsx_file, OPENPYXL_AVAILABLE
)

if OPENPYXL_AVAILABLE:
    import openpyxl


def build_workbook(rows, sheet_title="Plan", path=None):
    """Write a list of row tuples to a temporary .xlsx file and return its path."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    for row in rows:
        sheet.append(list(row))

    if path is None:
        handle = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        path = handle.name
        handle.close()

    workbook.save(path)
    return path


#: A plan shaped like a hand-built spreadsheet: title block, then the table.
SAMPLE_ROWS = [
    ('Implementation Plan',),
    (),
    ('ID', 'Phase', 'Task', 'Pred.', 'Duration (wd)', 'Start', 'End', 'Status'),
    (1, 'Phase One', 'Kick-off', '–', 5,
     datetime(2024, 1, 1), datetime(2024, 1, 5), 'Ongoing'),
    (2, 'Phase One', 'Analysis', '1', 5,
     datetime(2024, 1, 8), datetime(2024, 1, 12), 'Not started'),
    (3, 'Phase Two', 'Build', '1;2', 10,
     datetime(2024, 1, 15), datetime(2024, 1, 26), 'Done'),
]


@unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl is not installed")
class TestXLSXImporter(unittest.TestCase):
    """Tests for the XLSXImporter class."""

    def setUp(self):
        """Set up test fixtures."""
        self.importer = XLSXImporter()
        self.temp_files = []

    def tearDown(self):
        """Remove temporary workbooks."""
        for path in self.temp_files:
            if os.path.exists(path):
                os.unlink(path)

    def make(self, rows, sheet_title="Plan"):
        """Create a temporary workbook tracked for cleanup."""
        path = build_workbook(rows, sheet_title)
        self.temp_files.append(path)
        return path

    def test_import_basic_plan(self):
        """A header row below a title block is found and read."""
        project = self.importer.import_xlsx(self.make(SAMPLE_ROWS))

        self.assertIsNotNone(project)
        self.assertEqual(project.name, "Implementation Plan")

        names = [t.name for t in project.tasks]
        self.assertIn("Kick-off", names)
        self.assertIn("Analysis", names)
        self.assertIn("Build", names)

    def test_dates_are_read(self):
        """Start and end dates come through unchanged."""
        project = self.importer.import_xlsx(self.make(SAMPLE_ROWS))

        kickoff = project.get_task_by_id("1")
        self.assertEqual(kickoff.start_date, datetime(2024, 1, 1))
        self.assertEqual(kickoff.end_date, datetime(2024, 1, 5))
        self.assertEqual(kickoff.duration_days, 5)

    def test_multiple_predecessors(self):
        """A semicolon-separated predecessor cell yields several dependencies."""
        project = self.importer.import_xlsx(self.make(SAMPLE_ROWS))

        build = project.get_task_by_id("3")
        self.assertEqual(sorted(build.dependency_ids), ["1", "2"])

    def test_dash_means_no_predecessor(self):
        """An en-dash placeholder does not become a dependency."""
        project = self.importer.import_xlsx(self.make(SAMPLE_ROWS))

        kickoff = project.get_task_by_id("1")
        self.assertEqual(kickoff.dependency_ids, [])

    def test_status_maps_to_progress(self):
        """Status text is translated into a progress percentage."""
        project = self.importer.import_xlsx(self.make(SAMPLE_ROWS))

        self.assertEqual(project.get_task_by_id("1").progress, 50)   # Ongoing
        self.assertEqual(project.get_task_by_id("2").progress, 0)    # Not started
        self.assertEqual(project.get_task_by_id("3").progress, 100)  # Done

    def test_phases_become_parent_tasks(self):
        """Each phase becomes a parent task holding its rows as Subtasks."""
        project = self.importer.import_xlsx(self.make(SAMPLE_ROWS))

        roots = project.get_root_tasks()
        self.assertEqual([t.name for t in roots], ["Phase One", "Phase Two"])

        phase_one = roots[0]
        subtasks = project.get_subtasks(phase_one.id)
        self.assertEqual([t.name for t in subtasks], ["Kick-off", "Analysis"])
        for subtask in subtasks:
            self.assertEqual(subtask.task_type, "Subtask")

    def test_phase_parent_spans_its_children(self):
        """A phase parent covers the range of the rows inside it."""
        project = self.importer.import_xlsx(self.make(SAMPLE_ROWS))

        phase_one = project.get_root_tasks()[0]
        self.assertEqual(phase_one.start_date, datetime(2024, 1, 1))
        self.assertEqual(phase_one.end_date, datetime(2024, 1, 12))

    def test_grouping_can_be_disabled(self):
        """group_by_phase=False imports a flat task list."""
        importer = XLSXImporter(group_by_phase=False)
        project = importer.import_xlsx(self.make(SAMPLE_ROWS))

        self.assertEqual(len(project.tasks), 3)
        self.assertTrue(all(t.parent_task_id is None for t in project.tasks))

    def test_end_date_derived_from_working_days(self):
        """A 'Duration (wd)' column without an End column skips weekends."""
        rows = [
            ('ID', 'Task', 'Duration (wd)', 'Start'),
            (1, 'Week of work', 5, datetime(2024, 1, 1)),  # Mon
            (2, 'Spans a weekend', 10, datetime(2024, 1, 1)),
        ]
        project = self.importer.import_xlsx(self.make(rows))

        # 5 working days from Monday ends that Friday
        self.assertEqual(project.get_task_by_id("1").end_date, datetime(2024, 1, 5))
        # 10 working days ends on the Friday of the following week
        self.assertEqual(project.get_task_by_id("2").end_date, datetime(2024, 1, 12))

    def test_end_date_derived_from_calendar_days(self):
        """A plain 'Duration' column is treated as calendar days."""
        rows = [
            ('ID', 'Task', 'Duration', 'Start'),
            (1, 'Ten days', 10, datetime(2024, 1, 1)),
        ]
        project = self.importer.import_xlsx(self.make(rows))

        self.assertEqual(project.get_task_by_id("1").end_date, datetime(2024, 1, 10))

    def test_excel_serial_dates(self):
        """Dates stored as day serial numbers are converted."""
        rows = [
            ('ID', 'Task', 'Start', 'End'),
            (1, 'Serial dates', 45292, 45296),
        ]
        project = self.importer.import_xlsx(self.make(rows))

        task = project.get_task_by_id("1")
        self.assertEqual(task.start_date, datetime(2024, 1, 1))
        self.assertEqual(task.end_date, datetime(2024, 1, 5))

    def test_zero_duration_is_a_milestone(self):
        """A zero-duration row imports as a milestone with no end date."""
        rows = [
            ('ID', 'Task', 'Duration', 'Start'),
            (1, 'Go-Live', 0, datetime(2024, 1, 10)),
        ]
        project = self.importer.import_xlsx(self.make(rows))

        task = project.get_task_by_id("1")
        self.assertTrue(task.is_milestone)
        self.assertIsNone(task.end_date)

    def test_dependencies_by_name(self):
        """Predecessors given as task names resolve to the right tasks."""
        rows = [
            ('Task', 'Predecessors', 'Start', 'Duration'),
            ('First', '', datetime(2024, 1, 1), 3),
            ('Second', 'First', datetime(2024, 1, 4), 3),
        ]
        project = self.importer.import_xlsx(self.make(rows))

        first = next(t for t in project.tasks if t.name == 'First')
        second = next(t for t in project.tasks if t.name == 'Second')
        self.assertEqual(second.dependency_ids, [first.id])

    def test_unknown_predecessor_is_dropped(self):
        """A predecessor with no matching task does not create a dangling edge."""
        rows = [
            ('ID', 'Task', 'Pred.', 'Start', 'Duration'),
            (1, 'Only task', '99', datetime(2024, 1, 1), 3),
        ]
        project = self.importer.import_xlsx(self.make(rows))

        self.assertEqual(project.get_task_by_id("1").dependencies, [])

    def test_rows_below_a_blank_run_are_ignored(self):
        """Notes beneath the table are not read as tasks."""
        rows = [
            ('ID', 'Task', 'Start', 'Duration'),
            (1, 'Real task', datetime(2024, 1, 1), 3),
            (),
            (),
            (),
            (None, 'Legend: blue = planned', None, None),
        ]
        project = self.importer.import_xlsx(self.make(rows))

        self.assertEqual(len(project.tasks), 1)
        self.assertEqual(project.tasks[0].name, 'Real task')

    def test_hungarian_headers(self):
        """Accented Hungarian column headers are recognised."""
        rows = [
            ('Azonosító', 'Feladat', 'Kezdés', 'Befejezés', 'Státusz'),
            (1, 'Tesztelés', datetime(2024, 1, 1), datetime(2024, 1, 5), 'Folyamatban'),
        ]
        project = self.importer.import_xlsx(self.make(rows))

        self.assertEqual(len(project.tasks), 1)
        task = project.get_task_by_id("1")
        self.assertEqual(task.name, 'Tesztelés')
        self.assertEqual(task.start_date, datetime(2024, 1, 1))
        self.assertEqual(task.progress, 50)

    def test_explicit_parent_column_wins_over_phase(self):
        """A 'Parent Task' column is used instead of synthesising phases."""
        rows = [
            ('ID', 'Name', 'Parent Task', 'Start Date', 'End Date'),
            (1, 'Parent', None, datetime(2024, 1, 1), datetime(2024, 1, 10)),
            (2, 'Child', 'Parent', datetime(2024, 1, 1), datetime(2024, 1, 5)),
        ]
        project = self.importer.import_xlsx(self.make(rows))

        self.assertEqual(len(project.tasks), 2)
        child = project.get_task_by_id("2")
        self.assertEqual(child.parent_task_id, "1")
        self.assertEqual(child.task_type, "Subtask")

    def test_sheet_without_table_is_skipped(self):
        """A workbook whose first sheet has no table still imports."""
        workbook = openpyxl.Workbook()
        cover = workbook.active
        cover.title = "Cover"
        cover.append(["Some notes"])
        cover.append(["No table here"])

        data = workbook.create_sheet("Tasks")
        for row in [('ID', 'Task', 'Start', 'Duration'),
                    (1, 'Real task', datetime(2024, 1, 1), 3)]:
            data.append(list(row))

        handle = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        path = handle.name
        handle.close()
        self.temp_files.append(path)
        workbook.save(path)

        project = self.importer.import_xlsx(path)
        self.assertIsNotNone(project)
        self.assertEqual(len(project.tasks), 1)

    def test_import_nonexistent_file(self):
        """Importing a missing file returns None."""
        self.assertIsNone(self.importer.import_xlsx("/nonexistent/plan.xlsx"))

    def test_workbook_without_recognisable_table(self):
        """A workbook with no task table returns None."""
        rows = [
            ('Colour', 'Meaning'),
            ('Blue', 'Planned'),
        ]
        self.assertIsNone(self.importer.import_xlsx(self.make(rows)))

    def test_duplicate_ids_are_made_unique(self):
        """Two rows sharing an ID do not produce two tasks with one ID."""
        rows = [
            ('ID', 'Task', 'Pred.', 'Start', 'Duration'),
            (1, 'Alpha', '', datetime(2024, 1, 1), 3),
            (1, 'Beta', '', datetime(2024, 1, 5), 3),
            (2, 'Gamma', '1', datetime(2024, 1, 9), 3),
        ]
        project = self.importer.import_xlsx(self.make(rows))

        ids = [t.id for t in project.tasks]
        self.assertEqual(len(ids), len(set(ids)), f"duplicate ids: {ids}")

        # The first row keeps the sheet's ID, so its dependency still resolves
        alpha = next(t for t in project.tasks if t.name == 'Alpha')
        gamma = next(t for t in project.tasks if t.name == 'Gamma')
        self.assertEqual(gamma.dependency_ids, [alpha.id])

    def test_duplicate_id_is_reported(self):
        """A duplicated ID column value is logged rather than silently fixed."""
        from gantt_app.utils.log import setup_logging, get_log_text, reset_logging

        reset_logging()
        setup_logging(to_file=False, to_stderr=False)
        try:
            rows = [
                ('ID', 'Task', 'Start', 'Duration'),
                (1, 'Alpha', datetime(2024, 1, 1), 3),
                (1, 'Beta', datetime(2024, 1, 5), 3),
            ]
            self.importer.import_xlsx(self.make(rows))
            self.assertIn("Duplicate task ID", get_log_text())
        finally:
            reset_logging()

    def test_project_column_is_not_a_project_name_label(self):
        """A 'Project' column heading is not mistaken for a name label."""
        rows = [
            ('ID', 'Project', 'Task', 'Start', 'Duration'),
            (1, 'Apollo', 'Kick-off', datetime(2024, 1, 1), 5),
            (2, 'Apollo', 'Build', datetime(2024, 1, 8), 5),
        ]
        project = self.importer.import_xlsx(self.make(rows, "Portfolio"))

        # Falls back to the sheet name, not the neighbouring header cell
        self.assertEqual(project.name, "Portfolio")
        self.assertNotEqual(project.name, "Task")

    def test_project_name_label_is_still_read(self):
        """A real 'Project Name:' label is used as the project name."""
        rows = [
            ('Project Name:', 'Real Name'),
            (),
            ('ID', 'Task', 'Start', 'Duration'),
            (1, 'Kick-off', datetime(2024, 1, 1), 3),
        ]
        project = self.importer.import_xlsx(self.make(rows))

        self.assertEqual(project.name, "Real Name")

    def test_convenience_function(self):
        """import_xlsx_file works without constructing the importer."""
        project = import_xlsx_file(self.make(SAMPLE_ROWS))
        self.assertIsNotNone(project)
        self.assertEqual(project.name, "Implementation Plan")


@unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl is not installed")
class TestXLSXRoundTrip(unittest.TestCase):
    """Tests that an exported workbook can be imported back unchanged."""

    def setUp(self):
        """Import the sample plan and export it again."""
        from gantt_app.utils.xlsx_exporter import export_project_to_xlsx

        self.temp_files = []
        source = build_workbook(SAMPLE_ROWS)
        self.temp_files.append(source)

        self.original = import_xlsx_file(source)

        handle = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.exported_path = handle.name
        handle.close()
        self.temp_files.append(self.exported_path)

        self.assertTrue(export_project_to_xlsx(self.original, self.exported_path))
        self.reimported = import_xlsx_file(self.exported_path)

    def tearDown(self):
        """Remove temporary workbooks."""
        for path in self.temp_files:
            if os.path.exists(path):
                os.unlink(path)

    def test_reimport_succeeds(self):
        """An exported workbook is recognised by the importer."""
        self.assertIsNotNone(self.reimported)

    def test_task_count_preserved(self):
        """No tasks are gained or lost."""
        self.assertEqual(len(self.reimported.tasks), len(self.original.tasks))

    def test_project_name_preserved(self):
        """The project name survives via the Summary sheet."""
        self.assertEqual(self.reimported.name, self.original.name)

    def test_dates_and_progress_preserved(self):
        """Every task keeps its dates, progress and milestone flag."""
        original = {t.name: t for t in self.original.tasks}
        reimported = {t.name: t for t in self.reimported.tasks}

        self.assertEqual(set(original), set(reimported))
        for name, task in original.items():
            other = reimported[name]
            self.assertEqual(task.start_date, other.start_date, name)
            self.assertEqual(task.end_date, other.end_date, name)
            self.assertEqual(task.progress, other.progress, name)
            self.assertEqual(task.is_milestone, other.is_milestone, name)

    def test_dependencies_preserved(self):
        """Dependencies survive being written out as task names."""
        original = {t.name: t for t in self.original.tasks}
        reimported = {t.name: t for t in self.reimported.tasks}

        for name, task in original.items():
            expected = sorted(self.original.get_task_by_id(d).name
                              for d in task.dependency_ids)
            actual = sorted(self.reimported.get_task_by_id(d).name
                            for d in reimported[name].dependency_ids)
            self.assertEqual(expected, actual, name)

    def test_hierarchy_preserved(self):
        """Phase parents come back through the Parent Task column."""
        original = {t.name: t for t in self.original.tasks}
        reimported = {t.name: t for t in self.reimported.tasks}

        for name, task in original.items():
            expected = (self.original.get_task_by_id(task.parent_task_id).name
                        if task.parent_task_id else None)
            other = reimported[name]
            actual = (self.reimported.get_task_by_id(other.parent_task_id).name
                      if other.parent_task_id else None)
            self.assertEqual(expected, actual, name)

    def test_phases_are_not_nested_twice(self):
        """Re-importing does not wrap the phase parents in new parents."""
        self.assertEqual(len(self.reimported.get_root_tasks()),
                         len(self.original.get_root_tasks()))


@unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl is not installed")
class TestXLSXValueParsing(unittest.TestCase):
    """Tests for the importer's individual value coercions."""

    def setUp(self):
        """Set up test fixtures."""
        self.importer = XLSXImporter()

    def test_normalise_header_strips_accents_and_punctuation(self):
        """Header keys are lowercased, de-accented and stripped of colons."""
        self.assertEqual(self.importer._normalise_header("  Kezdés:  "), "kezdes")
        self.assertEqual(self.importer._normalise_header("Duration  (wd)"), "duration (wd)")

    def test_split_dependencies(self):
        """Predecessor cells split on common separators."""
        self.assertEqual(self.importer._split_dependencies("1;2"), ["1", "2"])
        self.assertEqual(self.importer._split_dependencies("1, 2 , 3"), ["1", "2", "3"])
        self.assertEqual(self.importer._split_dependencies("–"), [])
        self.assertEqual(self.importer._split_dependencies(None), [])
        self.assertEqual(self.importer._split_dependencies(4), ["4"])

    def test_split_dependencies_drops_lag_suffixes(self):
        """MS Project style lag notation is reduced to the task reference."""
        self.assertEqual(self.importer._split_dependencies("3FS+2d"), ["3"])
        self.assertEqual(self.importer._split_dependencies("7SS-1"), ["7"])
        self.assertEqual(self.importer._split_dependencies("12FF"), ["12"])

    def test_lag_stripping_leaves_task_names_intact(self):
        """A name containing the letters ss or ff is not truncated."""
        # 'process' ends in 'ss' and 'assessment' contains it
        self.assertEqual(
            self.importer._split_dependencies("Start the procurement demand process"),
            ["Start the procurement demand process"]
        )
        self.assertEqual(
            self.importer._split_dependencies("Functional risk assessment"),
            ["Functional risk assessment"]
        )
        self.assertEqual(self.importer._split_dependencies("Staff handoff"),
                         ["Staff handoff"])

    def test_slash_is_not_a_separator(self):
        """Task names routinely contain a slash, so it never splits."""
        self.assertEqual(self.importer._split_dependencies("Education / training"),
                         ["Education / training"])

    def test_dependency_name_containing_a_comma(self):
        """A single name with a comma resolves as a whole before splitting."""
        rows = [
            ('Task', 'Predecessors', 'Start', 'Duration'),
            ('Analysis, phase 2', '', datetime(2024, 1, 1), 3),
            ('Build', 'Analysis, phase 2', datetime(2024, 1, 4), 3),
        ]
        handle = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        path = handle.name
        handle.close()
        try:
            build_workbook(rows, path=path)
            project = self.importer.import_xlsx(path)

            first = next(t for t in project.tasks if t.name == 'Analysis, phase 2')
            build = next(t for t in project.tasks if t.name == 'Build')
            self.assertEqual(build.dependency_ids, [first.id])
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_parse_cell_date_formats(self):
        """Dates parse from datetimes, serials and common strings."""
        self.assertEqual(
            self.importer._parse_cell_date(datetime(2024, 1, 1)),
            datetime(2024, 1, 1)
        )
        self.assertEqual(
            self.importer._parse_cell_date(45292), datetime(2024, 1, 1)
        )
        self.assertEqual(
            self.importer._parse_cell_date("2024-01-01"), datetime(2024, 1, 1)
        )
        self.assertIsNone(self.importer._parse_cell_date("not a date"))
        self.assertIsNone(self.importer._parse_cell_date(None))

    def test_end_date_for_a_working_day_duration_skips_weekends(self):
        """A duration counted in working days steps over the weekend."""
        monday = datetime(2024, 1, 1)
        self.assertEqual(self.importer._end_date_for(monday, 5, True),
                         datetime(2024, 1, 5))     # Mon to Fri
        self.assertEqual(self.importer._end_date_for(monday, 6, True),
                         datetime(2024, 1, 8))     # spills to the Monday
        self.assertEqual(self.importer._end_date_for(monday, 1, True), monday)

    def test_end_date_for_a_calendar_duration_does_not(self):
        """A sheet counting calendar days is taken at its word."""
        monday = datetime(2024, 1, 1)
        self.assertEqual(self.importer._end_date_for(monday, 7, False),
                         datetime(2024, 1, 7))     # straight through Sunday

    def test_start_date_for_works_backwards_from_the_end(self):
        """A sheet giving an end and a duration gets a working-day start."""
        friday = datetime(2024, 1, 5)
        self.assertEqual(self.importer._start_date_for(friday, 5, True),
                         datetime(2024, 1, 1))     # back to the Monday
        self.assertEqual(self.importer._start_date_for(friday, 5, False),
                         datetime(2024, 1, 1))

    def test_progress_from_percentage_column(self):
        """An explicit progress column takes precedence over status."""
        self.assertEqual(
            self.importer._progress_from_row({'progress': 75, 'status': 'Done'}),
            75
        )
        self.assertEqual(self.importer._progress_from_row({'progress': 0.5}), 50)
        self.assertEqual(self.importer._progress_from_row({'progress': 150}), 100)

    def test_progress_defaults_to_zero(self):
        """An unrecognised status leaves progress at zero."""
        self.assertEqual(self.importer._progress_from_row({'status': 'Whatever'}), 0)
        self.assertEqual(self.importer._progress_from_row({}), 0)


if __name__ == '__main__':
    unittest.main()
