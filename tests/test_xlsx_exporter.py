"""
Tests for the XLSX export: the plan sheet the exporter writes.

WHY THIS MODULE EXISTS:
======================
The export used to be three sheets of raw fields, which was a faithful record
of the model and no use to anybody who wanted to look at the plan. What it
writes now is a project plan sheet - a title, an editable start date, one row
per piece of work grouped by phase, and a week-by-week chart drawn in the
cells beside it - and the things worth pinning down are the ones a reader
would notice: where the table starts, what the columns are, and whether the
formulas in it agree with the plan they came from.

DEVELOPMENT NOTES:
------------------
The sheet is live: Start and End are WORKDAY formulas over an editable
duration. A formula that disagreed with the dates this application worked out
would open in Excel showing a plan nobody scheduled, so several of these
tests are about exactly that - the formula is written where it reproduces the
answer and the date itself is written where it would not.

Nothing here needs a display.
"""

import os
import tempfile
import unittest
from datetime import datetime, timedelta

from gantt_app.models import Project, Task
from gantt_app.utils.xlsx_exporter import (
    OPENPYXL_AVAILABLE, export_project_to_xlsx, generate_xlsx_bytes,
)

if OPENPYXL_AVAILABLE:
    import openpyxl


@unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl is not installed")
class ExporterTestCase(unittest.TestCase):
    """A plan with a phase, a deliverable, work inside it and a milestone."""

    def setUp(self):
        """Build the plan and write it out."""
        self.temp_files = []
        self.project = self.build_project()
        self.sheet, self.workbook = self.export(self.project)

    def tearDown(self):
        """Remove the workbooks."""
        for path in self.temp_files:
            if os.path.exists(path):
                os.unlink(path)

    def build_project(self):
        """A small plan starting on Monday 6 July 2026."""
        project = Project(name="Tosca Implementation")
        base = datetime(2026, 7, 6)

        project.add_task(Task(id="P1", name="1. Procurement", task_type="Phase",
                              start_date=base, end_date=base))
        project.add_task(Task(id="D1", name="Signed contract",
                              task_type="Deliverable", parent_task_id="P1",
                              start_date=base, end_date=base))
        project.add_task(Task(id="T1", name="Business case", task_type="Subtask",
                              parent_task_id="D1", start_date=base,
                              end_date=base + timedelta(days=4), progress=50))
        project.add_task(Task(id="T2", name="Procurement demand",
                              task_type="Subtask", parent_task_id="D1",
                              start_date=base, end_date=base + timedelta(days=9)))
        project.get_task_by_id("T2").add_dependency("T1", "FS", "Hard")

        project.add_task(Task(id="P2", name="2. Requirements", task_type="Phase",
                              start_date=base, end_date=base))
        project.add_task(Task(id="T3", name="URS", task_type="Subtask",
                              parent_task_id="P2", start_date=base,
                              end_date=base + timedelta(days=9), progress=100))
        project.get_task_by_id("T3").add_dependency("T2", "FS", "Hard")

        project.reschedule()
        return project

    def export(self, project):
        """Write the project out and hand back its plan sheet."""
        handle = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        path = handle.name
        handle.close()
        self.temp_files.append(path)

        self.assertTrue(export_project_to_xlsx(project, path))
        workbook = openpyxl.load_workbook(path)
        return workbook.worksheets[0], workbook

    def row_of(self, name):
        """The sheet row a named task was written to."""
        for row in range(6, self.sheet.max_row + 1):
            if self.sheet.cell(row=row, column=3).value == name:
                return row
        self.fail(f"no row for {name!r}")

    def column(self, row, letter):
        """One cell by row and column letter."""
        return self.sheet[f"{letter}{row}"].value


class TestTheShapeOfTheSheet(ExporterTestCase):
    """Where the fixed parts of the plan sheet sit."""

    def test_the_sheet_is_named_after_the_project(self):
        """A reader opening the file sees the plan, not 'Tasks'."""
        self.assertEqual(self.sheet.title, "Tosca Implementation")

    def test_the_title_names_the_plan(self):
        """The first row says what this is."""
        self.assertIn("Tosca Implementation", str(self.sheet['A1'].value))

    def test_the_start_date_is_one_editable_cell(self):
        """
        The whole plan hangs off it.

        Written as a DATE() call rather than a date value so that changing it
        keeps it a date, which is what the formulas below it need.
        """
        self.assertEqual(self.sheet['A3'].value, "Project Start Date:")
        self.assertEqual(self.sheet['C3'].value, "=DATE(2026,7,6)")

    def test_the_headings_are_the_plan_columns(self):
        """The ten fields, in the order a plan is read in."""
        headings = [self.sheet.cell(row=5, column=c).value
                    for c in range(1, 11)]

        self.assertEqual(headings, [
            'ID', 'Phase', 'Task', 'Responsible (A)', 'Key Deliverable',
            'Pred.', 'Duration (wd)', 'Start', 'End', 'Status',
        ])

    def test_the_header_stays_put_when_the_plan_scrolls(self):
        """Everything above the first task is frozen."""
        self.assertEqual(self.sheet.freeze_panes, "A6")

    def test_the_totals_are_on_their_own_sheet(self):
        """The plan sheet is the plan; the numbers sit behind it."""
        self.assertIn("Summary", self.workbook.sheetnames)
        summary = self.workbook["Summary"]
        labels = {summary.cell(row=r, column=1).value: summary.cell(row=r, column=2).value
                  for r in range(1, summary.max_row + 1)}
        self.assertEqual(labels["Project Name:"], "Tosca Implementation")


class TestWhichTasksGetRows(ExporterTestCase):
    """The rows are the work, and the grouping is a column."""

    def test_the_work_gets_the_rows(self):
        """Every leaf task is a row."""
        names = [self.sheet.cell(row=r, column=3).value
                 for r in range(6, self.sheet.max_row + 1)]

        self.assertEqual(names, ["Business case", "Procurement demand", "URS"])

    def test_a_phase_is_a_column_not_a_row(self):
        """
        A Phase brackets other rows rather than being work of its own.

        This layout has one grouping column, so that is where the phase goes;
        giving it a row as well would have it appear twice.
        """
        names = [self.sheet.cell(row=r, column=3).value
                 for r in range(6, self.sheet.max_row + 1)]

        self.assertNotIn("1. Procurement", names)
        self.assertEqual(self.column(self.row_of("Business case"), 'B'),
                         "1. Procurement")

    def test_a_deliverable_is_named_beside_its_work(self):
        """The level the Phase column cannot show is still readable."""
        self.assertEqual(self.column(self.row_of("Business case"), 'E'),
                         "Signed contract")

    def test_each_phase_gets_its_own_banding(self):
        """Two phases are told apart down the sheet by colour."""
        first = self.sheet[f"B{self.row_of('Business case')}"].fill.fgColor.rgb
        second = self.sheet[f"B{self.row_of('URS')}"].fill.fgColor.rgb

        self.assertNotEqual(first, second)

    def test_an_empty_phase_still_gets_a_row(self):
        """Work that has not been broken down yet is still work."""
        project = Project(name="Just a phase")
        project.add_task(Task(id="P", name="Discovery", task_type="Phase",
                              start_date=datetime(2026, 7, 6),
                              end_date=datetime(2026, 7, 10)))
        sheet, _ = self.export(project)

        self.assertEqual(sheet.cell(row=6, column=3).value, "Discovery")


class TestTheSheetIsLive(ExporterTestCase):
    """Duration is editable and the dates follow it."""

    def test_the_first_task_hangs_off_the_start_date(self):
        """Changing one cell moves the plan."""
        self.assertEqual(self.column(self.row_of("Business case"), 'H'), "=$C$3")

    def test_a_following_task_chains_onto_its_predecessor(self):
        """The next working day after the row it follows finishes."""
        row = self.row_of("Business case")

        self.assertEqual(self.column(self.row_of("Procurement demand"), 'H'),
                         f"=WORKDAY(I{row},1)")

    def test_the_finish_is_the_duration_walked_over_the_calendar(self):
        """WORKDAY, so a task crossing a weekend reaches further out."""
        row = self.row_of("Procurement demand")

        self.assertEqual(self.column(row, 'I'), f"=WORKDAY(H{row},G{row}-1)")

    def test_the_duration_is_the_working_days_the_task_holds(self):
        """Not the calendar span: this one crosses a weekend."""
        task = self.project.get_task_by_id("T1")

        self.assertEqual(self.column(self.row_of("Business case"), 'G'),
                         self.project.working_duration(task))

    def test_the_timeline_is_drawn_from_the_dates(self):
        """
        Each week asks whether it overlaps the task, so the bars follow an
        edit rather than being fixed marks.
        """
        row = self.row_of("Business case")

        self.assertEqual(
            self.column(row, 'K'),
            f'=IF(AND(K$5<=$I{row},K$5+6>=$H{row}),"█","")'
        )

    def test_the_weeks_run_from_the_start_date(self):
        """The chart's first column is the project start, then weekly."""
        self.assertEqual(self.sheet['K5'].value, "=$C$3")
        self.assertEqual(self.sheet['L5'].value, "=K5+7")


class TestFormulasNeverDisagreeWithThePlan(ExporterTestCase):
    """A live sheet is only worth having while it says what the plan says."""

    def test_a_start_start_link_is_written_as_a_date(self):
        """
        A WORKDAY chain can only say Finish-Start.

        Writing one for a task held by a Start-Start link would open in Excel
        on a date this application never scheduled, so the date itself is
        written and the link is left as information in the Pred. column.

        The Start-Start is put on the third task, which starts partway through
        the plan. On the second it would land on the project start date, and
        pointing at the start cell reproduces that exactly - so the formula
        would be right and this would prove nothing.
        """
        project = self.build_project()
        third = project.get_task_by_id("T3")
        third.dependencies = []
        third.add_dependency("T2", "SS", "Hard")
        project.apply_dependency_constraints(third)
        project.reschedule()

        sheet, _ = self.export(project)
        row = next(r for r in range(6, sheet.max_row + 1)
                   if sheet.cell(row=r, column=3).value == "URS")

        self.assertEqual(sheet[f"H{row}"].value, third.start_date)
        self.assertNotEqual(third.start_date, project.start_date)
        self.assertEqual(sheet[f"F{row}"].value, "2SS")

    def test_a_lagged_link_is_written_as_a_date(self):
        """The chain cannot count the days a lag adds, so it is not used."""
        project = self.build_project()
        second = project.get_task_by_id("T2")
        second.dependencies = []
        second.add_dependency("T1", "FS", "Hard", lag=3)
        project.apply_dependency_constraints(second)
        project.reschedule()

        sheet, _ = self.export(project)
        row = next(r for r in range(6, sheet.max_row + 1)
                   if sheet.cell(row=r, column=3).value == "Procurement demand")

        self.assertEqual(sheet[f"H{row}"].value, second.start_date)
        self.assertEqual(sheet[f"F{row}"].value, "1FS+3")

    def test_an_unlinked_task_that_starts_later_keeps_its_date(self):
        """Only a task that starts the project can point at the start cell."""
        project = Project(name="Two starts")
        project.add_task(Task(id="A", name="Early",
                              start_date=datetime(2026, 7, 6),
                              end_date=datetime(2026, 7, 10)))
        project.add_task(Task(id="B", name="Later",
                              start_date=datetime(2026, 8, 3),
                              end_date=datetime(2026, 8, 7)))
        project.reschedule()

        sheet, _ = self.export(project)

        self.assertEqual(sheet["H6"].value, "=$C$3")
        self.assertEqual(sheet["H7"].value, datetime(2026, 8, 3))

    def test_no_predecessor_reads_as_a_dash(self):
        """An empty cell would read as an unanswered question."""
        self.assertEqual(self.column(self.row_of("Business case"), 'F'), "–")


class TestMilestonesAndStatus(ExporterTestCase):
    """The two things a row says about itself besides its dates."""

    def test_a_milestone_takes_no_time(self):
        """Nought days, finishing the day it starts."""
        project = self.build_project()
        project.add_task(Task(id="M", name="Go-Live", task_type="Milestone",
                              parent_task_id="P2",
                              start_date=datetime(2026, 8, 3)))
        project.reschedule()

        sheet, _ = self.export(project)
        row = next(r for r in range(6, sheet.max_row + 1)
                   if sheet.cell(row=r, column=3).value == "Go-Live")

        self.assertEqual(sheet[f"G{row}"].value, 0)
        self.assertEqual(sheet[f"I{row}"].value, f"=$H{row}")

    def test_progress_becomes_a_status_word(self):
        """A percentage means nothing on a printed plan."""
        self.assertEqual(self.column(self.row_of("Business case"), 'J'),
                         "Ongoing")
        self.assertEqual(self.column(self.row_of("Procurement demand"), 'J'),
                         "Not started")
        self.assertEqual(self.column(self.row_of("URS"), 'J'), "Done")


class TestHolidaysReachTheFormulas(ExporterTestCase):
    """WORKDAY skips weekends on its own; the rest has to be given to it."""

    def test_no_holiday_sheet_when_the_calendar_has_none(self):
        """A plain plan gets a plain WORKDAY."""
        self.assertNotIn("Holidays", self.workbook.sheetnames)
        row = self.row_of("Procurement demand")
        self.assertEqual(self.column(row, 'I'), f"=WORKDAY(H{row},G{row}-1)")

    def test_a_calendar_with_holidays_writes_them_out(self):
        """
        The dates go on a hidden sheet WORKDAY is pointed at.

        Without them the exported sheet would recalculate onto dates the
        application does not schedule - the one thing a live sheet must not
        do.
        """
        project = self.build_project()
        project.set_holiday_countries(["HU"])

        sheet, workbook = self.export(project)

        self.assertIn("Holidays", workbook.sheetnames)
        self.assertEqual(workbook["Holidays"].sheet_state, 'hidden')

        row = next(r for r in range(6, sheet.max_row + 1)
                   if sheet.cell(row=r, column=3).value == "Procurement demand")
        self.assertEqual(sheet[f"I{row}"].value,
                         f"=WORKDAY(H{row},G{row}-1,Holidays!$A:$A)")


@unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl is not installed")
class TestTheOtherEntryPoint(unittest.TestCase):
    """generate_xlsx_bytes, for a caller that does not want a file."""

    def test_it_returns_a_workbook(self):
        """The bytes open as the same plan sheet."""
        from io import BytesIO

        project = Project(name="In memory")
        project.add_task(Task(id="A", name="Work",
                              start_date=datetime(2026, 7, 6),
                              end_date=datetime(2026, 7, 10)))

        data = generate_xlsx_bytes(project)

        self.assertIsNotNone(data)
        workbook = openpyxl.load_workbook(BytesIO(data))
        self.assertEqual(workbook.worksheets[0].cell(row=6, column=3).value,
                         "Work")


if __name__ == '__main__':
    unittest.main()
