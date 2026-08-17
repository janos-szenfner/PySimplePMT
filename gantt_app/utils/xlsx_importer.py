"""
Excel XLSX importer for the Gantt Project Management Tool.

Reads spreadsheet-based project plans and converts them into Project objects.

DEVELOPMENT NOTES:
------------------
There is no single "XLSX project plan" format, so this importer is driven by
column headers rather than fixed positions. It locates the header row by
scoring candidate rows against a table of known column aliases, which lets it
read both hand-built plans (ID / Phase / Task / Pred. / Duration / Start / End
/ Status) and files produced by this application's own xlsx_exporter.

Two shapes of hierarchy are supported:
  * an explicit 'Parent Task' column (what xlsx_exporter writes), and
  * a 'Phase'/'Section' column, whose distinct values become parent tasks so
    the plan's grouping survives the import.

Excel stores dates either as real datetimes or as day serial numbers; both are
handled. Formula-driven date columns are read from their cached values, so the
workbook must have been saved by a spreadsheet application rather than
generated without a calculation pass.
"""

import re
from datetime import datetime, date, timedelta
from typing import Optional, List, Dict, Any, Tuple
from pathlib import Path

from gantt_app.models import Project, Task
from gantt_app.workdaycalendar import WorkingCalendar
from gantt_app.utils.log import get_logger

logger = get_logger(__name__)

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


#: Excel's day-zero. Excel treats 1900 as a leap year, and using 1899-12-30 as
#: the epoch absorbs that off-by-one for every date after 1900-03-01.
EXCEL_EPOCH = datetime(1899, 12, 30)


class XLSXImporter:
    """
    Imports Excel (.xlsx) project plans and converts them to Project objects.
    """

    #: Normalised header text -> logical column name.
    COLUMN_ALIASES: Dict[str, str] = {
        # Identifier
        'id': 'id', 'task id': 'id', '#': 'id', 'no': 'id', 'nr': 'id',
        'sorszam': 'id', 'azonosito': 'id',
        # Name
        'name': 'name', 'task': 'name', 'task name': 'name',
        'activity': 'name', 'description': 'name',
        'feladat': 'name', 'megnevezes': 'name', 'tevekenyseg': 'name',
        # Grouping
        'phase': 'phase', 'section': 'phase', 'group': 'phase',
        'stage': 'phase', 'workstream': 'phase',
        'fazis': 'phase', 'szakasz': 'phase', 'csoport': 'phase',
        # Explicit parent (written by xlsx_exporter)
        'parent task': 'parent', 'parent': 'parent', 'parent name': 'parent',
        'szulo': 'parent',
        # Dates
        'start': 'start', 'start date': 'start', 'begin': 'start',
        'begin date': 'start', 'kezdes': 'start', 'kezdet': 'start',
        'kezdo datum': 'start',
        'end': 'end', 'end date': 'end', 'finish': 'end',
        'finish date': 'end', 'due': 'end', 'due date': 'end',
        'befejezes': 'end', 'veg': 'end', 'zaro datum': 'end',
        # Duration
        'duration': 'duration', 'duration (wd)': 'duration',
        'duration (days)': 'duration', 'duration (d)': 'duration',
        'days': 'duration', 'working days': 'duration', 'effort': 'duration',
        'idotartam': 'duration', 'tartam': 'duration', 'munkanap': 'duration',
        # Dependencies
        'pred': 'dependencies', 'pred.': 'dependencies',
        'predecessor': 'dependencies', 'predecessors': 'dependencies',
        'dependency': 'dependencies', 'dependencies': 'dependencies',
        'depends on': 'dependencies', 'elozmeny': 'dependencies',
        'megelozo': 'dependencies',
        # Progress / status
        'progress': 'progress', 'progress (%)': 'progress',
        '% complete': 'progress', 'percent complete': 'progress',
        'completion': 'progress', 'complete': 'progress',
        'keszultseg': 'progress',
        'status': 'status', 'statusz': 'status', 'allapot': 'status',
        # Flags
        'milestone': 'milestone', 'merfoldko': 'milestone',
        'type': 'type', 'tipus': 'type',
        'color': 'color', 'colour': 'color', 'szin': 'color',
    }

    #: Status text -> progress percentage.
    STATUS_PROGRESS: Dict[str, int] = {
        'done': 100, 'complete': 100, 'completed': 100, 'finished': 100,
        'closed': 100, 'kesz': 100, 'befejezett': 100, 'lezart': 100,
        'ongoing': 50, 'in progress': 50, 'in-progress': 50, 'started': 50,
        'wip': 50, 'active': 50, 'folyamatban': 50, 'elkezdett': 50,
        'not started': 0, 'notstarted': 0, 'pending': 0, 'planned': 0,
        'new': 0, 'open': 0, 'todo': 0, 'nem kezdodott el': 0,
        'nem indult': 0, 'tervezett': 0,
    }

    #: Cell values that mean "no predecessor".
    EMPTY_TOKENS = {'', '-', '--', '–', '—', 'n/a', 'na', 'none', 'nincs'}

    #: A header row must match at least this many known columns.
    MIN_HEADER_MATCHES = 3

    #: How far down the sheet to look for a header row.
    HEADER_SEARCH_ROWS = 30

    #: Blank rows tolerated before the task table is considered finished.
    MAX_BLANK_RUN = 3

    def __init__(self, group_by_phase: bool = True):
        """
        PARAMETERS:
        -----------
        group_by_phase : bool, optional
            When True (default), each distinct value of the Phase column
            becomes a parent Task and its rows become Sub-Tasks of it. Ignored
            when the sheet carries an explicit 'Parent Task' column.
        """
        self.default_color = "#1f6aa5"
        self.milestone_color = "#e74c3c"
        self.phase_color = "#34495e"
        self.group_by_phase = group_by_phase

    # ------------------------------------------------------------------
    # Value coercion
    # ------------------------------------------------------------------

    def _normalise_header(self, value: Any) -> str:
        """
        Reduce a header cell to a comparable key.

        Lowercases, strips accents that Hungarian headers commonly carry,
        collapses whitespace and drops a trailing colon.
        """
        if value is None:
            return ''

        text = str(value).strip().lower().rstrip(':').strip()
        text = re.sub(r'\s+', ' ', text)

        accents = {
            'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ö': 'o', 'ő': 'o',
            'ú': 'u', 'ü': 'u', 'ű': 'u',
        }
        for accented, plain in accents.items():
            text = text.replace(accented, plain)

        return text

    def _parse_cell_date(self, value: Any) -> Optional[datetime]:
        """Coerce a cell value into a datetime, or None if it is not a date."""
        if value is None or value == '':
            return None

        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day)

        # Excel day serial
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            try:
                return EXCEL_EPOCH + timedelta(days=float(value))
            except (OverflowError, ValueError):
                return None

        text = str(value).strip()
        if not text:
            return None

        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%m/%d/%Y',
                    '%d.%m.%Y', '%Y.%m.%d', '%Y. %m. %d.', '%d-%m-%Y',
                    '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S'):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue

        return None

    def _parse_number(self, value: Any) -> Optional[float]:
        """Coerce a cell value into a number, or None."""
        if value is None or isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            return float(value)

        match = re.search(r'-?\d+(?:[.,]\d+)?', str(value))
        if not match:
            return None
        try:
            return float(match.group(0).replace(',', '.'))
        except ValueError:
            return None

    def _parse_bool(self, value: Any) -> bool:
        """Coerce a cell value into a boolean flag."""
        if isinstance(value, bool):
            return value
        if value is None:
            return False
        if isinstance(value, (int, float)):
            return value != 0
        return str(value).strip().lower() in {
            'yes', 'y', 'true', 't', 'x', '1', 'igen', 'milestone'
        }

    def _progress_from_row(self, values: Dict[str, Any]) -> int:
        """
        Derive a 0-100 progress figure from a Progress and/or Status column.

        An explicit Progress column wins; a Status column is mapped through
        STATUS_PROGRESS. Fractions between 0 and 1 are read as percentages.
        """
        raw = values.get('progress')
        number = self._parse_number(raw)
        if number is not None:
            if 0 < number <= 1 and isinstance(raw, float):
                number *= 100
            return max(0, min(100, int(round(number))))

        status = values.get('status')
        if status is not None:
            key = self._normalise_header(status)
            if key in self.STATUS_PROGRESS:
                return self.STATUS_PROGRESS[key]

        return 0

    #: The calendar a stated duration is counted against: Monday to Friday.
    #:
    #: Public holidays are not modelled. The spreadsheets this reads use
    #: Excel's WORKDAY without a holiday list, so weekends alone reproduce
    #: them, and it is the same calendar the application schedules on - see
    #: gantt_app.workdaycalendar.
    CALENDAR = WorkingCalendar()

    def _end_date_for(self, start: datetime, duration: int,
                      working_days: bool) -> datetime:
        """
        The inclusive end date of a task the sheet gave a duration for.

        PARAMETERS:
        -----------
        start : datetime
            The task's start date.
        duration : int
            Length in days, the start day included.
        working_days : bool
            Whether the sheet counts that duration in working days.
        """
        if working_days:
            return self.CALENDAR.add_working_days(start, duration)
        return start + timedelta(days=duration - 1)

    def _start_date_for(self, end: datetime, duration: int,
                        working_days: bool) -> datetime:
        """The inclusive start date, for a sheet that gives the end instead."""
        if working_days:
            return self.CALENDAR.subtract_working_days(end, duration)
        return end - timedelta(days=duration - 1)

    #: An MS Project style reference: a task number, an optional dependency
    #: type, and an optional lag - "12", "3FS", "3FS+2d", "7SS-1".
    LAG_REFERENCE = re.compile(
        r'^(\d+)\s*(?:FS|SS|FF|SF)\s*(?:[+-]\s*\d+\s*\w*)?$',
        re.IGNORECASE
    )

    def _split_dependencies(self, value: Any) -> List[str]:
        """
        Split a predecessor cell into individual references.

        DEVELOPMENT NOTES:
        ------------------
        Only ';' , ',' and '|' separate references. '/' is deliberately not a
        separator because task names routinely contain one ("Education /
        training"), and a dependency written as a name has to survive intact.
        A lag suffix is stripped only from tokens that are entirely a numeric
        reference plus a dependency type - matching it loosely truncated any
        name containing the letters ss or ff, such as "process".
        """
        if value is None:
            return []

        if isinstance(value, (int, float)) and not isinstance(value, bool):
            if float(value).is_integer():
                return [str(int(value))]
            return [str(value)]

        tokens = re.split(r'[;,|]+', str(value))
        result = []
        for token in tokens:
            token = token.strip()
            if not token or token.lower() in self.EMPTY_TOKENS:
                continue
            lag_match = self.LAG_REFERENCE.match(token)
            if lag_match:
                token = lag_match.group(1)
            result.append(token)
        return result

    # ------------------------------------------------------------------
    # Sheet inspection
    # ------------------------------------------------------------------

    def _map_header_row(self, row: Tuple) -> Dict[str, int]:
        """
        Map a candidate header row to {logical column name: column index}.

        The first occurrence of each logical column wins, so a stray repeat
        further right cannot displace the real one.
        """
        mapping: Dict[str, int] = {}
        for index, cell in enumerate(row):
            key = self._normalise_header(cell)
            if not key:
                continue
            logical = self.COLUMN_ALIASES.get(key)
            if logical and logical not in mapping:
                mapping[logical] = index
        return mapping

    def _find_header_row(self, rows: List[Tuple]) -> Optional[Tuple[int, Dict[str, int]]]:
        """
        Locate the header row and its column mapping.

        RETURNS:
        --------
        Optional[Tuple[int, Dict[str, int]]]
            (row index, column mapping), or None when no row looks like a
            header. A usable header needs a name column plus enough other
            recognised columns to be more than a coincidence.
        """
        best: Optional[Tuple[int, Dict[str, int]]] = None

        for index, row in enumerate(rows[:self.HEADER_SEARCH_ROWS]):
            mapping = self._map_header_row(row)
            if 'name' not in mapping:
                continue
            if len(mapping) < self.MIN_HEADER_MATCHES:
                continue
            if not ({'start', 'duration', 'end'} & set(mapping)):
                continue
            if best is None or len(mapping) > len(best[1]):
                best = (index, mapping)

        return best

    def _select_sheet(self, workbook) -> Optional[Tuple[Any, int, Dict[str, int]]]:
        """
        Pick the worksheet holding the task table.

        RETURNS:
        --------
        Optional[Tuple[worksheet, int, Dict[str, int]]]
            The sheet with the richest recognised header row, along with that
            row's index and column mapping.
        """
        best = None

        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            found = self._find_header_row(rows)
            if found is None:
                continue
            header_index, mapping = found
            if best is None or len(mapping) > len(best[2]):
                best = (sheet, header_index, mapping)

        return best

    # ------------------------------------------------------------------
    # Row -> Task
    # ------------------------------------------------------------------

    def _read_rows(self, rows: List[Tuple], header_index: int,
                   mapping: Dict[str, int]) -> List[Dict[str, Any]]:
        """
        Read the task rows below the header into plain dictionaries.

        Stops after a run of blank rows so trailing notes or legends beneath
        the table are not mistaken for tasks.
        """
        records: List[Dict[str, Any]] = []
        blank_run = 0

        for row in rows[header_index + 1:]:
            values = {
                logical: (row[index] if index < len(row) else None)
                for logical, index in mapping.items()
            }

            name = values.get('name')
            if name is None or str(name).strip() == '':
                blank_run += 1
                if blank_run >= self.MAX_BLANK_RUN:
                    break
                continue

            blank_run = 0
            values['name'] = str(name).strip()
            records.append(values)

        return records

    def _build_task(self, values: Dict[str, Any], row_number: int,
                    duration_in_working_days: bool,
                    used_ids: Optional[set] = None) -> Optional[Task]:
        """
        Convert one spreadsheet record into a Task.

        PARAMETERS:
        -----------
        values : Dict[str, Any]
            The record read from one spreadsheet row.
        row_number : int
            The 1-based sheet row, used to name rows with no ID.
        duration_in_working_days : bool
            Whether the duration column counts working days.
        used_ids : Optional[set]
            IDs already taken. Passing this makes the returned task's ID
            unique and records it.

        DEVELOPMENT NOTES:
        ------------------
        Task IDs must be unique: dependencies, parent references, the
        treeview and Project.remove_task all address tasks by ID. A
        hand-maintained sheet can easily repeat a value in the ID column, so
        a duplicate falls back to the row-based identifier rather than
        producing two tasks that answer to the same ID.
        """
        name = values['name']

        raw_id = values.get('id')
        if raw_id is None or str(raw_id).strip() == '':
            task_id = f"row_{row_number}"
        elif isinstance(raw_id, float) and raw_id.is_integer():
            task_id = str(int(raw_id))
        else:
            task_id = str(raw_id).strip()

        if used_ids is not None:
            if task_id in used_ids:
                logger.warning(
                    "Duplicate task ID %r on row %d (%r); using %r instead",
                    task_id, row_number, name, f"row_{row_number}"
                )
                task_id = f"row_{row_number}"
                suffix = 2
                while task_id in used_ids:
                    task_id = f"row_{row_number}_{suffix}"
                    suffix += 1
            used_ids.add(task_id)

        start_date = self._parse_cell_date(values.get('start'))
        end_date = self._parse_cell_date(values.get('end'))
        duration = self._parse_number(values.get('duration'))

        # A plan may give start+end, start+duration, or end+duration
        if start_date is None and end_date is not None and duration:
            start_date = self._start_date_for(end_date, int(duration),
                                             duration_in_working_days)
        if start_date is None:
            return None
        if end_date is None and duration is not None:
            span = int(duration)
            if span > 0:
                end_date = self._end_date_for(start_date, span,
                                              duration_in_working_days)

        is_milestone = self._parse_bool(values.get('milestone'))
        if not is_milestone and duration is not None and duration == 0:
            is_milestone = True
        if not is_milestone and values.get('type') is not None:
            is_milestone = 'milestone' in self._normalise_header(values['type'])

        colour = values.get('color')
        colour = str(colour).strip() if colour else None
        if not colour or not re.match(r'^#[0-9a-fA-F]{6}$', colour):
            colour = self.milestone_color if is_milestone else self.default_color

        return Task(
            id=task_id,
            name=name,
            start_date=start_date,
            end_date=None if is_milestone else end_date,
            progress=self._progress_from_row(values),
            dependencies=[],
            color=colour,
            is_milestone=is_milestone,
            task_type="Task",
            parent_task_id=None
        )

    def _resolve_dependencies(self, records: List[Dict[str, Any]],
                              tasks: List[Task]) -> None:
        """
        Turn predecessor references into task IDs.

        References may be given as the sheet's own ID values or as task names;
        both are resolved, and anything unmatched is dropped rather than
        producing a dangling edge.
        """
        by_id = {task.id: task for task in tasks}
        by_name = {task.name.strip().lower(): task for task in tasks}

        def resolve(token: str) -> Optional[Task]:
            """Match a single reference against task IDs then task names."""
            target = by_id.get(token) or by_name.get(token.strip().lower())
            if target is None and re.match(r'^\d+(\.0)?$', token):
                target = by_id.get(str(int(float(token))))
            return target

        for record, task in zip(records, tasks):
            raw = record.get('dependencies')
            if raw is None or str(raw).strip() == '':
                continue

            # Try the whole cell as one name first, so a task name containing
            # a separator ("Analysis, phase 2") is not torn apart
            tokens = [str(raw).strip()]
            if resolve(tokens[0]) is None:
                tokens = self._split_dependencies(raw)

            for token in tokens:
                target = resolve(token)
                if target is not None and target.id != task.id:
                    task.add_dependency(target.id)

    def _apply_explicit_parents(self, records: List[Dict[str, Any]],
                                tasks: List[Task]) -> None:
        """Wire up a 'Parent Task' column that names an existing task."""
        by_id = {task.id: task for task in tasks}
        by_name = {task.name.strip().lower(): task for task in tasks}

        for record, task in zip(records, tasks):
            raw = record.get('parent')
            if raw is None or str(raw).strip() == '':
                continue
            token = str(raw).strip()
            parent = by_id.get(token) or by_name.get(token.lower())
            if parent is not None and parent.id != task.id:
                task.parent_task_id = parent.id
                task.task_type = "Subtask"

    def _group_by_phase(self, records: List[Dict[str, Any]],
                        tasks: List[Task]) -> List[Task]:
        """
        Turn the Phase column into parent tasks holding their rows as Sub-Tasks.

        RETURNS:
        --------
        List[Task]
            The full task list with each phase's parent inserted immediately
            before the rows belonging to it. Rows with no phase stay at root.
        """
        members: Dict[str, List[Task]] = {}
        order: List[str] = []

        for record, task in zip(records, tasks):
            raw = record.get('phase')
            if raw is None or str(raw).strip() == '':
                continue
            phase = str(raw).strip()
            if phase not in members:
                members[phase] = []
                order.append(phase)
            members[phase].append(task)

        if not members:
            return tasks

        used_ids = {task.id for task in tasks}
        phase_tasks: Dict[str, Task] = {}

        for phase in order:
            group = members[phase]
            starts = [t.start_date for t in group if t.start_date]
            ends = [t.end_date for t in group if t.end_date]
            ends.extend(t.start_date for t in group
                        if t.end_date is None and t.start_date)
            if not starts:
                continue

            base = re.sub(r'[^a-zA-Z0-9_]', '_', phase.lower()).strip('_')
            phase_id = f"phase_{base}" if base else "phase"
            counter = 2
            while phase_id in used_ids:
                phase_id = f"phase_{base}_{counter}"
                counter += 1
            used_ids.add(phase_id)

            parent = Task(
                id=phase_id,
                name=phase,
                start_date=min(starts),
                end_date=max(ends) if ends else None,
                progress=0,
                dependencies=[],
                color=self.phase_color,
                is_milestone=False,
                task_type="Task",
                parent_task_id=None
            )
            phase_tasks[phase] = parent

            for member in group:
                member.task_type = "Subtask"
                member.parent_task_id = phase_id

        ordered: List[Task] = []
        emitted = set()

        for record, task in zip(records, tasks):
            raw = record.get('phase')
            phase = str(raw).strip() if raw is not None else ''
            if phase in phase_tasks and phase not in emitted:
                ordered.append(phase_tasks[phase])
                emitted.add(phase)
            ordered.append(task)

        return ordered

    #: Labels that introduce the project name on a summary sheet.
    PROJECT_NAME_LABELS = {'project name', 'project', 'projekt neve', 'projekt'}

    def _is_label_cell(self, value: Any) -> bool:
        """
        Check whether a cell reads as a label rather than a column header.

        DEVELOPMENT NOTES:
        ------------------
        A trailing colon is what separates 'Project Name:' sitting beside its
        value from a 'Project' column heading a table of values below it.
        xlsx_exporter writes the colon, so requiring it costs nothing and
        stops a portfolio-style sheet with a Project column from having its
        neighbouring header read as the project name.
        """
        return value is not None and str(value).strip().endswith(':')

    def _project_name_from_labels(self, workbook) -> Optional[str]:
        """
        Look for a 'Project Name:' label and return the value beside it.

        RETURNS:
        --------
        Optional[str]
            The value in the first cell to the right of the label, or None
            when no such label exists.

        DEVELOPMENT NOTES:
        ------------------
        This is what xlsx_exporter writes onto its Summary sheet, so finding
        it lets an exported workbook round-trip with its name intact instead
        of falling back to the 'Tasks' worksheet title.

        Two guards keep a *column* called 'Project' from being mistaken for
        this label: the cell must end with a colon, and the value taken from
        beside it must not itself be a recognised column heading.
        """
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(max_row=30, values_only=True):
                for index, cell in enumerate(row):
                    if not self._is_label_cell(cell):
                        continue
                    if self._normalise_header(cell) not in self.PROJECT_NAME_LABELS:
                        continue
                    for candidate in row[index + 1:]:
                        if candidate is None or not str(candidate).strip():
                            continue
                        # A neighbouring column heading is not a project name
                        if self._normalise_header(candidate) in self.COLUMN_ALIASES:
                            break
                        return str(candidate).strip()
        return None

    def _extract_project_name(self, sheet, rows: List[Tuple],
                              header_index: int) -> str:
        """
        Work out a project name.

        Prefers a title-looking cell above the table, and falls back to the
        worksheet name.
        """
        for row in rows[:header_index]:
            for cell in row:
                if cell is None:
                    continue
                text = str(cell).strip()
                if len(text) < 3:
                    continue
                # Skip label cells such as "Project Start Date:"
                if text.endswith(':'):
                    continue
                if self._normalise_header(text) in self.COLUMN_ALIASES:
                    continue
                return text

        return sheet.title or "Imported Excel Project"

    # ------------------------------------------------------------------
    # Entry point
    # ------------------------------------------------------------------

    def import_xlsx(self, filepath: str) -> Optional[Project]:
        """
        Import an .xlsx file and convert it to a Project object.

        PARAMETERS:
        -----------
        filepath : str
            Path to the .xlsx workbook.

        RETURNS:
        --------
        Optional[Project]
            The imported project, or None if the file is missing, unreadable,
            or contains no recognisable task table.
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("Error: openpyxl library is required for Excel import")
            return None

        try:
            if not Path(filepath).exists():
                logger.warning(f"File not found: {filepath}")
                return None

            workbook = openpyxl.load_workbook(filepath, data_only=True)

            selected = self._select_sheet(workbook)
            if selected is None:
                logger.error("Error: no recognisable task table found in workbook")
                return None

            sheet, header_index, mapping = selected
            rows = list(sheet.iter_rows(values_only=True))

            duration_header = ''
            if 'duration' in mapping:
                header_row = rows[header_index]
                if mapping['duration'] < len(header_row):
                    duration_header = self._normalise_header(
                        header_row[mapping['duration']]
                    )
            duration_in_working_days = (
                'wd' in duration_header or 'working' in duration_header
                or 'munkanap' in duration_header
            )

            records = self._read_rows(rows, header_index, mapping)
            if not records:
                logger.error("Error: task table contains no rows")
                return None

            tasks: List[Task] = []
            kept_records: List[Dict[str, Any]] = []
            used_ids: set = set()

            for offset, record in enumerate(records):
                task = self._build_task(
                    record,
                    header_index + 2 + offset,
                    duration_in_working_days,
                    used_ids
                )
                if task is not None:
                    tasks.append(task)
                    kept_records.append(record)

            if not tasks:
                logger.error("Error: no rows carried a usable date")
                return None

            self._resolve_dependencies(kept_records, tasks)

            if 'parent' in mapping:
                self._apply_explicit_parents(kept_records, tasks)
            elif self.group_by_phase and 'phase' in mapping:
                tasks = self._group_by_phase(kept_records, tasks)

            project_name = (self._project_name_from_labels(workbook)
                            or self._extract_project_name(sheet, rows, header_index))
            project = Project(name=project_name, tasks=tasks)
            logger.info("Imported %d task(s) from XLSX file %s",
                        len(project.tasks), filepath)
            return project

        except Exception as e:
            logger.exception(f"Error importing XLSX file: {e}")
            return None


def import_xlsx_file(filepath: str) -> Optional[Project]:
    """Import an .xlsx file and return a Project object."""
    importer = XLSXImporter()
    return importer.import_xlsx(filepath)
