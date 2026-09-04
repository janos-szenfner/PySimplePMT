"""
XLSX export: the project as a plan sheet somebody can work in.

WHY THIS MODULE EXISTS:
======================
A spreadsheet is where a plan gets circulated, argued over and edited by
people who do not have this application. What it exports therefore has to be
a plan, not a database dump. The first version wrote three sheets of raw
fields - Tasks, Dependencies, Summary - which was a faithful record of the
model and no use at all to anybody who wanted to look at the plan.

This writes the layout project plans are actually kept in: a title, an
editable project start date, one row per piece of work grouped by phase, and
a week-by-week bar chart drawn in the cells to the right.

WHAT MAKES IT A SPREADSHEET RATHER THAN A PICTURE:
==================================================
The sheet is live. Duration is a number the reader can change; Start and End
are WORKDAY formulas over it, so re-planning in Excel behaves the way
re-planning here does - weekends are skipped and a task pushed out drags the
chain behind it. The timeline bars are formulas over Start and End, so they
follow. Changing the start date in one cell moves the whole plan.

A formula is only written where it reproduces the date this application
already worked out; see _start_formula. Anywhere it would not - a task with
no predecessor, or one held by a Start-Start or Finish-Finish link, neither
of which a WORKDAY chain can express - the real date is written instead. A
sheet that is live but wrong would be worse than one that is merely static.

DEVELOPMENT NOTES:
------------------
Rows are the leaves of the plan: the work. A row with children is a
bracket over other rows rather than work of its own, so it appears as the
Phase column beside its work and as the colour banding down the sheet, which
is how the format expresses grouping. Nesting deeper than that is flattened -
the layout has one grouping column - and the Key Deliverable column says
what the row itself produces, taken from its notes.

Optional Dependency: needs openpyxl, and says so rather than failing
obscurely when it is missing.
"""

import os
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, TYPE_CHECKING

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None  # type: ignore

from gantt_app.models import Project, Task
from gantt_app.utils.log import get_logger
from gantt_app.workdaycalendar import (
    DEFAULT_NON_WORKING_DAYS, WorkingCalendar,
)

logger = get_logger(__name__)

if TYPE_CHECKING:
    if OPENPYXL_AVAILABLE:
        from openpyxl import Workbook as WorkbookType
    else:
        WorkbookType = Any  # type: ignore


# ---------------------------------------------------------------------------
# The look of the sheet
# ---------------------------------------------------------------------------

#: Every colour the sheet uses, as Excel ARGB.
HEADER_BG = 'FF1F4E79'          # the dark blue header band
HEADER_TEXT = 'FFFFFFFF'
TITLE_TEXT = 'FF1F4E79'
NOTE_TEXT = 'FF808080'
EDITABLE_BG = 'FFFFF2CC'        # the pale yellow of a cell meant to be typed in
EDITABLE_TEXT = 'FF0000FF'      # and its blue text, the spreadsheet convention
BAR_TEXT = 'FF2E75B6'           # the timeline bars

#: Banding down the Phase column, one colour per phase, cycled.
PHASE_FILLS = ('FFDDEBF7', 'FFE2EFDA', 'FFFCE4D6', 'FFEDEDED', 'FFFFF2CC',
               'FFE4DFEC')

#: Status wording, and the fill that goes with it.
STATUS_NOT_STARTED = 'Not started'
STATUS_ONGOING = 'Ongoing'
STATUS_DONE = 'Done'
STATUS_FILLS = {
    STATUS_NOT_STARTED: 'FFFFF2CC',
    STATUS_ONGOING: 'FFC6EFCE',
    STATUS_DONE: 'FFA9D08E',
}

#: The character a timeline cell draws a covered week with.
BAR_GLYPH = '█'

FONT_NAME = 'Arial'

#: Where the fixed part of the sheet sits.
ROW_TITLE = 1
ROW_START_DATE = 3
ROW_HEADER = 5
ROW_FIRST_TASK = 6

#: Column letters for the fields, and the first timeline column.
COL_ID, COL_PHASE, COL_TASK, COL_OWNER = 'A', 'B', 'C', 'D'
COL_DELIVERABLE, COL_PRED, COL_DURATION = 'E', 'F', 'G'
COL_START, COL_END, COL_STATUS = 'H', 'I', 'J'
FIRST_TIMELINE_COL = 11         # K

#: The header row's captions, in order from column A.
HEADINGS = ('ID', 'Phase', 'Task', 'Responsible (A)', 'Key Deliverable',
            'Pred.', 'Duration (wd)', 'Start', 'End', 'Status')

#: Column widths, by letter. The timeline columns share FIRST_TIMELINE_COL's.
COLUMN_WIDTHS = {
    'A': 4, 'B': 32, 'C': 42, 'D': 26, 'E': 46, 'F': 8,
    'G': 10, 'H': 11, 'I': 11, 'J': 12,
}
TIMELINE_WIDTH = 3.66

#: How many weeks of timeline to draw, whatever the plan is. A short plan
#: still wants a chart, and a very long one is capped so the sheet does not
#: become tens of thousands of cells of formula.
MIN_TIMELINE_WEEKS = 8
MAX_TIMELINE_WEEKS = 104

DATE_FORMAT = 'yyyy\\-mm\\-dd'
WEEK_FORMAT = 'mm\\-dd'

#: What goes in the Pred. column when there is nothing to point at.
NO_PREDECESSOR = '–'

#: Name of the sheet holding the holiday dates, when the project has any.
HOLIDAY_SHEET = 'Holidays'


def _thin_border():
    """The hairline box every cell in the table wears."""
    side = Side(style='thin')
    return Border(left=side, right=side, top=side, bottom=side)


def _sheet_title(project: Project) -> str:
    """
    A worksheet name Excel will accept for this project.

    Excel forbids []:*?/\\ in a sheet name and caps it at 31 characters, and
    it refuses to open a file that breaks either rule.
    """
    name = (project.name or 'Project Plan').strip() or 'Project Plan'
    for bad in '[]:*?/\\':
        name = name.replace(bad, ' ')
    return name[:31].strip() or 'Project Plan'


# ---------------------------------------------------------------------------
# Working out what goes in the rows
# ---------------------------------------------------------------------------

def _plan_rows(project: Project) -> List[Task]:
    """
    The tasks that get a row: the work, in the order the plan holds it.

    RETURNS:
    --------
    List[Task]
        Every task with nothing hanging off it. A row with children and
        children is a bracket over other rows rather than work of its own, so
        it becomes the Phase column and the colour banding instead of a row.
        One with nothing inside it is work that has not been broken down yet,
        and does get a row.
    """
    with_children = project.get_summary_task_ids()
    return [task for task in project.tasks if task.id not in with_children]


def _ancestors(project: Project, task: Task) -> List[Task]:
    """
    A task's parents, nearest first, guarded against a broken parent chain.
    """
    found: List[Task] = []
    seen = {task.id}
    current = task
    while current.parent_task_id and current.parent_task_id not in seen:
        parent = project.get_task_by_id(current.parent_task_id)
        if parent is None:
            break
        seen.add(parent.id)
        found.append(parent)
        current = parent
    return found


def _phase_of(project: Project, task: Task) -> str:
    """
    The name to put in the Phase column.

    The nearest ancestor that is a Phase, or failing that the outermost
    ancestor of any kind - a plan grouped into plain Tasks still wants its
    grouping shown. A task at the top level has no phase.
    """
    ancestors = _ancestors(project, task)
    for ancestor in ancestors:
        if ancestor.task_type == 'Phase':
            return ancestor.name
    return ancestors[-1].name if ancestors else ''


def _deliverable_of(project: Project, task: Task) -> str:
    """
    The name to put in the Key Deliverable column.

    The first line of the task's own notes, which is where a plan says what
    a row produces.

    DEVELOPMENT NOTES:
    ------------------
    This used to name the nearest ancestor of the Deliverable type, falling
    back to the notes where a plan had none. That type is no longer offered -
    a plan runs Phase, Task, Subtask - so the fallback is the whole rule now.

    The column keeps its heading. What a piece of work delivers is something
    a reader of a project spreadsheet looks for, and it does not stop being
    one because no row is called that any more.
    """
    return (task.details or '').strip().splitlines()[0] if task.details else ''


def _status_of(task: Task) -> str:
    """
    Which of the three states a task is in, for its colour.

    The word alone, without the percentage - the fill is keyed by it, and a
    key that changed with every task's progress would need a colour per
    percentage.

    DEVELOPMENT NOTES:
    ------------------
    This is not Task.status, which says whether a row is Active, Estimated
    or Inactive and is the planner's own answer. The spreadsheet's Status
    column is read
    off the progress and says how far the work has got - two different
    questions that the word Status is used for on both sides of the export.
    Neither is derived from the other, and this one does not read the field.
    """
    if task.progress >= 100:
        return STATUS_DONE
    if task.progress > 0:
        return STATUS_ONGOING
    return STATUS_NOT_STARTED


def _status_text(task: Task) -> str:
    """
    What the Status cell says.

    RETURNS:
    --------
    str
        "Ongoing - 30%" for work that has started and is not finished, and
        the bare word for the other two.

    DEVELOPMENT NOTES:
    ------------------
    A plan that says only "Ongoing" says the least useful thing it could:
    every task between the first day and the last is ongoing, and which of
    them is nearly done is the question a reader of a status column is
    actually asking. Not started and Done carry no percentage because theirs
    is implied - nought and a hundred - and printing it would be noise.

    The importer reads the number back out; see XLSXImporter's
    _progress_from_row.
    """
    status = _status_of(task)
    if status == STATUS_ONGOING:
        return f"{status} - {int(task.progress)}%"
    return status


def _predecessor_text(task: Task, numbers: Dict[str, int]) -> str:
    """
    The Pred. cell: the row numbers this task follows.

    PARAMETERS:
    -----------
    task : Task
        The dependent task.
    numbers : Dict[str, int]
        Task ID to the number in the ID column.

    RETURNS:
    --------
    str
        Numbers joined with ';', each carrying its link type and lag where
        those are not the plain Finish-Start with no lag that a bare number
        means - "4", "4SS", "4FS+2". That is MS Project's notation, which is
        what a reader of a plan expects and what the importer reads back.

    DEVELOPMENT NOTES:
    ------------------
    A predecessor that has no row of its own - a summary, which this layout
    does not give rows to - is left out. Pointing at a row number that is not
    in the sheet would be worse than saying nothing.
    """
    parts = []
    for dependency in task.dependencies:
        number = numbers.get(dependency.task_id)
        if number is None:
            continue
        text = str(number)
        # The type is spelt out whenever anything follows it. A bare "1+3"
        # is not the notation - the reader cannot tell the lag from the
        # reference, and the importer's own reference pattern wants the type
        # before the sign, so the link came back pointing at nothing.
        if dependency.dep_type != 'FS' or dependency.lag:
            text += dependency.dep_type
        if dependency.lag:
            text += f"{dependency.lag:+d}"
        parts.append(text)
    return ';'.join(parts) if parts else NO_PREDECESSOR


def _followed_rows(task: Task, rows: Dict[str, int]) -> List[int]:
    """
    The sheet rows a task follows by a plain Finish-Start link.

    RETURNS:
    --------
    List[int]
        Row numbers, empty when the task has no such link. Only these can be
        expressed as a WORKDAY chain: a Start-Start or Finish-Finish link
        places the task somewhere the formula has no way to say, and a lag
        would need the formula to count days the chain does not.
    """
    followed = []
    for dependency in task.dependencies:
        if dependency.dep_type != 'FS' or dependency.lag:
            return []
        row = rows.get(dependency.task_id)
        if row is None:
            return []
        followed.append(row)
    return followed


def _workday(expression: str, days: str, holidays: bool) -> str:
    """WORKDAY over the project's holiday list, when it has one."""
    if holidays:
        return f"WORKDAY({expression},{days},{HOLIDAY_SHEET}!$A:$A)"
    return f"WORKDAY({expression},{days})"


def _excel_calendar(holiday_dates: List[date]) -> WorkingCalendar:
    """
    The calendar Excel will actually use, which is not always the plan's.

    RETURNS:
    --------
    WorkingCalendar
        Saturday and Sunday off, plus the dates written to the holiday sheet.

    DEVELOPMENT NOTES:
    ------------------
    WORKDAY's week is fixed: it skips Saturday and Sunday and takes a list of
    extra dates to skip, and there is nothing else it can be told. So a
    project working an unusual week, or - the case this was written for -
    carrying a manual override that puts a Saturday to work, is scheduled by
    this application onto dates no WORKDAY formula can reproduce.

    Building that restricted calendar explicitly is what lets the writers
    below tell the difference. Comparing against the project's own calendar
    would compare the plan with itself and agree every time, which is how a
    worked Saturday came to be written as a live formula that recalculated,
    in Excel, onto a Monday.
    """
    return WorkingCalendar(non_working_days=DEFAULT_NON_WORKING_DAYS,
                           holidays=set(holiday_dates))


def _end_formula_agrees(task: Task, project: Project,
                        excel: WorkingCalendar) -> bool:
    """
    Whether WORKDAY over the start and duration lands on the task's finish.

    False for a task whose span covers a day the two calendars disagree
    about - a Saturday the plan works and WORKDAY will not - in which case
    the finish is written as a date rather than a formula that would open
    showing a different one.
    """
    if task.start_date is None or task.end_date is None:
        return False

    duration = max(project.working_duration(task), 1)
    return (excel.add_working_days(task.start_date.date(), duration)
            == task.end_date.date())


def _start_formula(task: Task, project: Project, rows: Dict[str, int],
                   holidays: bool, excel: WorkingCalendar) -> Optional[str]:
    """
    The Start cell's formula, or None to write the date itself.

    RETURNS:
    --------
    Optional[str]
        A WORKDAY expression chaining this task onto the rows it follows, or
        '=$C$3' for a task that starts the project. None when no formula
        would land on the date the task actually has.

    DEVELOPMENT NOTES:
    ------------------
    The formula is checked against the answer before it is written. A live
    sheet is worth having, but only while it agrees with the plan it came
    from: a WORKDAY chain over the wrong links would open in Excel showing
    dates this application never scheduled, and the reader has no way to tell
    which of the two is the plan.

    So the arithmetic is done here first, with the calendar WORKDAY actually
    implements - see _excel_calendar, which is not the project's own once a
    manual override puts a weekend day to work - and the formula is written
    only where the two agree.
    """
    calendar = excel
    followed = _followed_rows(task, rows)

    if not followed:
        if project.start_date is None:
            return None
        if task.start_date.date() != project.start_date.date():
            return None
        return "=$C$3"

    predecessors = [project.get_task_by_id(d.task_id) for d in task.dependencies]
    finishes = [(p.end_date or p.start_date) for p in predecessors if p]
    if not finishes:
        return None

    expected = calendar.get_next_working_day(max(finishes) + timedelta(days=1))
    if expected.date() != task.start_date.date():
        return None

    cells = ','.join(f"{COL_END}{row}" for row in sorted(followed))
    inner = cells if len(followed) == 1 else f"MAX({cells})"
    return "=" + _workday(inner, "1", holidays)


def _timeline_weeks(project: Project) -> int:
    """How many weekly columns the chart needs to cover the plan."""
    if project.start_date is None or project.end_date is None:
        return MIN_TIMELINE_WEEKS
    days = max((project.end_date - project.start_date).days, 0)
    weeks = days // 7 + 2
    return max(MIN_TIMELINE_WEEKS, min(weeks, MAX_TIMELINE_WEEKS))


def _project_holidays(project: Project) -> List[date]:
    """
    Every non-working date the plan spans that is not a weekend.

    RETURNS:
    --------
    List[date]
        Sorted holidays inside the plan's span, for the sheet WORKDAY is
        pointed at. Empty when the calendar declares none, in which case no
        holiday sheet is written and the formulas skip weekends only.

    DEVELOPMENT NOTES:
    ------------------
    A manual override that takes a working day off is a holiday as far as
    WORKDAY is concerned, and is picked up by the walk below like any other.
    One that puts a *weekend* day to work is not expressible here at all -
    WORKDAY always skips Saturday and Sunday, and no holiday list can tell it
    otherwise. That case needs nothing doing about it: _start_formula checks
    every formula against the date this application actually scheduled and
    writes the literal date wherever the two disagree, so the sheet stays
    truthful and only loses the live recalculation on those rows.
    """
    calendar = project.calendar
    if not (calendar.holidays or calendar.recurring_holidays
            or calendar.countries or calendar.overrides):
        return []

    if project.start_date is None or project.end_date is None:
        return []

    found = []
    current = project.start_date.date()
    last = project.end_date.date() + timedelta(days=30)
    while current <= last:
        if (current.weekday() not in calendar.non_working_days
                and not calendar.is_working_day(current)):
            found.append(current)
        current += timedelta(days=1)
    return found


# ---------------------------------------------------------------------------
# Drawing the sheet
# ---------------------------------------------------------------------------

def _write_title(sheet, project: Project) -> None:
    """The name of the plan, and the start date box that drives it."""
    title = sheet.cell(row=ROW_TITLE, column=1,
                       value=f"{project.name or 'Project'} – Project Plan")
    title.font = Font(name=FONT_NAME, size=14, bold=True, color=TITLE_TEXT)
    sheet.row_dimensions[ROW_TITLE].height = 18

    label = sheet.cell(row=ROW_START_DATE, column=1, value="Project Start Date:")
    label.font = Font(name=FONT_NAME, size=10, bold=True)

    start = project.start_date or datetime.now()
    box = sheet[f"{COL_TASK}{ROW_START_DATE}"]
    box.value = f"=DATE({start.year},{start.month},{start.day})"
    box.font = Font(name=FONT_NAME, size=10, color=EDITABLE_TEXT)
    box.fill = PatternFill('solid', fgColor=EDITABLE_BG)
    box.border = _thin_border()
    box.number_format = DATE_FORMAT

    note = sheet.cell(row=ROW_START_DATE, column=4,
                      value="◄ Editable (working days follow automatically "
                            "via WORKDAY)")
    note.font = Font(name=FONT_NAME, size=9, color=NOTE_TEXT)


def _write_header(sheet, weeks: int) -> None:
    """The field captions, then a column per week of the plan."""
    fill = PatternFill('solid', fgColor=HEADER_BG)
    border = _thin_border()

    for index, caption in enumerate(HEADINGS, start=1):
        cell = sheet.cell(row=ROW_HEADER, column=index, value=caption)
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=HEADER_TEXT)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)

    for offset in range(weeks):
        column = FIRST_TIMELINE_COL + offset
        cell = sheet.cell(row=ROW_HEADER, column=column)
        if offset == 0:
            cell.value = f"=${COL_TASK}${ROW_START_DATE}"
        else:
            previous = get_column_letter(column - 1)
            cell.value = f"={previous}{ROW_HEADER}+7"
        cell.font = Font(name=FONT_NAME, size=8, bold=True, color=HEADER_TEXT)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        cell.number_format = WEEK_FORMAT

    sheet.row_dimensions[ROW_HEADER].height = 90


def _write_task_row(sheet, row: int, task: Task, project: Project,
                    number: int, numbers: Dict[str, int],
                    rows: Dict[str, int], phase_fill: Optional[str],
                    weeks: int, holidays: bool,
                    excel: WorkingCalendar) -> None:
    """One piece of work, across the fields and the timeline."""
    border = _thin_border()
    body = Font(name=FONT_NAME, size=10)
    centred = Alignment(horizontal='center')
    wrapped = Alignment(vertical='center', wrap_text=True)

    identifier = sheet.cell(row=row, column=1, value=number)
    identifier.font = body
    identifier.alignment = centred
    identifier.border = border

    phase = sheet.cell(row=row, column=2, value=_phase_of(project, task))
    phase.font = Font(name=FONT_NAME, size=9)
    phase.alignment = Alignment(vertical='center', wrap_text=True)
    phase.border = border
    if phase_fill:
        phase.fill = PatternFill('solid', fgColor=phase_fill)

    for column, value in ((3, task.name),
                          (4, ''),                      # Responsible: the
                                                        # model has no owner,
                                                        # so the column is
                                                        # there to be filled in
                          (5, _deliverable_of(project, task))):
        cell = sheet.cell(row=row, column=column, value=value)
        cell.font = body
        cell.alignment = wrapped
        cell.border = border

    predecessors = sheet.cell(row=row, column=6,
                              value=_predecessor_text(task, numbers))
    predecessors.font = body
    predecessors.alignment = centred
    predecessors.border = border

    # A milestone marks a moment and takes no time, which a duration of
    # nought days is how this format says. Everything else holds at least a
    # day of work.
    length = 0 if task.effective_milestone else max(
        project.working_duration(task), 1)
    duration = sheet.cell(row=row, column=7, value=length)
    duration.font = Font(name=FONT_NAME, size=10, color=EDITABLE_TEXT)
    duration.fill = PatternFill('solid', fgColor=EDITABLE_BG)
    duration.alignment = centred
    duration.border = border

    start = sheet.cell(row=row, column=8)
    formula = _start_formula(task, project, rows, holidays, excel)
    start.value = formula if formula else task.start_date
    start.font = body
    start.alignment = centred
    start.border = border
    start.number_format = DATE_FORMAT

    end = sheet.cell(row=row, column=9)
    if task.effective_milestone:
        # No length, so it ends the day it starts
        end.value = f"=${COL_START}{row}"
    elif formula and _end_formula_agrees(task, project, excel):
        end.value = "=" + _workday(f"{COL_START}{row}", f"{COL_DURATION}{row}-1",
                                   holidays)
    else:
        # Either the start is a date already, or WORKDAY would walk this
        # task's span onto a different finish - see _end_formula_agrees.
        end.value = task.end_date or task.start_date
    end.font = body
    end.alignment = centred
    end.border = border
    end.number_format = DATE_FORMAT

    status = sheet.cell(row=row, column=10, value=_status_text(task))
    status.font = Font(name=FONT_NAME, size=10, color=EDITABLE_TEXT)
    status.fill = PatternFill('solid', fgColor=STATUS_FILLS[_status_of(task)])
    status.alignment = Alignment(vertical='center')
    status.border = border

    _write_timeline(sheet, row, weeks)
    sheet.row_dimensions[row].height = 30


def _write_timeline(sheet, row: int, weeks: int) -> None:
    """
    The bar, as one formula per week.

    DEVELOPMENT NOTES:
    ------------------
    Each cell asks whether its week overlaps the task at all - the week starts
    on or before the finish, and ends on or after the start - and draws a
    block if it does. Written as formulas over the Start and End cells rather
    than as fixed marks, so editing a duration redraws the chart.
    """
    border = _thin_border()
    font = Font(name=FONT_NAME, size=10, color=BAR_TEXT)
    centred = Alignment(horizontal='center')

    for offset in range(weeks):
        column = get_column_letter(FIRST_TIMELINE_COL + offset)
        cell = sheet.cell(row=row, column=FIRST_TIMELINE_COL + offset)
        cell.value = (
            f'=IF(AND({column}${ROW_HEADER}<=${COL_END}{row},'
            f'{column}${ROW_HEADER}+6>=${COL_START}{row}),"{BAR_GLYPH}","")'
        )
        cell.font = font
        cell.alignment = centred
        cell.border = border


def _write_holiday_sheet(workbook, holidays: List[date]) -> None:
    """
    The dates WORKDAY is told to skip beyond the weekend.

    Kept on its own sheet because that is what WORKDAY's third argument
    wants - a range - and hidden, because it is machinery rather than plan.
    """
    sheet = workbook.create_sheet(HOLIDAY_SHEET)
    sheet.cell(row=1, column=1, value="Non-working dates")
    sheet.cell(row=1, column=1).font = Font(name=FONT_NAME, bold=True)
    for index, day in enumerate(holidays, start=2):
        cell = sheet.cell(row=index, column=1, value=day)
        cell.number_format = DATE_FORMAT
    sheet.column_dimensions['A'].width = 18
    sheet.sheet_state = 'hidden'


def _write_summary_sheet(workbook, project: Project) -> None:
    """
    The project's totals, for whoever wants the numbers rather than the plan.

    The project name lives here too, which is how the importer recovers it -
    the plan sheet's title cell is prose and the sheet's own name is capped
    at 31 characters.
    """
    sheet = workbook.create_sheet("Summary")
    milestones = sum(1 for t in project.tasks if t.effective_milestone)
    containers = sum(1 for t in project.tasks if t.is_container)

    rows = [
        ("Project Summary", ""),
        ("Project Name:", project.name),
        ("Total Tasks:", len(project.tasks)),
        ("Phases and groupings:", containers),
        ("Milestones:", milestones),
        ("Start Date:", project.start_date.strftime('%Y-%m-%d')
         if project.start_date else ''),
        ("End Date:", project.end_date.strftime('%Y-%m-%d')
         if project.end_date else ''),
        ("Completed:", sum(1 for t in project.tasks if t.progress >= 100)),
        ("In Progress:", sum(1 for t in project.tasks
                             if 0 < t.progress < 100)),
        ("Not Started:", sum(1 for t in project.tasks if t.progress == 0)),
    ]
    for index, (label, value) in enumerate(rows, start=1):
        caption = sheet.cell(row=index, column=1, value=label)
        caption.font = Font(name=FONT_NAME, bold=index == 1)
        sheet.cell(row=index, column=2, value=value)

    sheet.column_dimensions['A'].width = 26
    sheet.column_dimensions['B'].width = 30


def _create_tasks_workbook(project: Project):
    """
    Build the workbook: the plan sheet, and the totals behind it.

    RETURNS:
    --------
    Workbook
        An openpyxl workbook ready to save.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _sheet_title(project)

    tasks = _plan_rows(project)
    weeks = _timeline_weeks(project)
    holidays = _project_holidays(project)
    excel = _excel_calendar(holidays)

    _write_title(sheet, project)
    _write_header(sheet, weeks)

    # Numbered and placed before anything is written, so a row can point at
    # one further down the sheet - which a plan whose links run backwards up
    # the list will do.
    #
    # The ID column carries the number the application shows beside the row,
    # not a count of the rows in this sheet. The two are not the same: a
    # sheet holds one row per piece of work and shows the phases as a column
    # beside them, so the numbers here have gaps where the phases sit. That
    # is the right way round - somebody reading the sheet against the plan is
    # looking for the task the plan calls 4, and a sheet that called it 3
    # would be quietly wrong. See Project.display_ids.
    shown = project.display_ids()
    numbers = {task.id: shown[task.id] for task in tasks if task.id in shown}
    rows = {task.id: ROW_FIRST_TASK + index
            for index, task in enumerate(tasks)}

    phase_colours: Dict[str, str] = {}
    for index, task in enumerate(tasks):
        phase = _phase_of(project, task)
        if phase and phase not in phase_colours:
            phase_colours[phase] = PHASE_FILLS[len(phase_colours)
                                               % len(PHASE_FILLS)]
        _write_task_row(sheet, rows[task.id], task, project,
                        numbers[task.id], numbers, rows,
                        phase_colours.get(phase), weeks, bool(holidays),
                        excel)

    for letter, width in COLUMN_WIDTHS.items():
        sheet.column_dimensions[letter].width = width
    for offset in range(weeks):
        letter = get_column_letter(FIRST_TIMELINE_COL + offset)
        sheet.column_dimensions[letter].width = TIMELINE_WIDTH

    # Everything above the first task stays put while the plan scrolls
    sheet.freeze_panes = f"A{ROW_FIRST_TASK}"

    if holidays:
        _write_holiday_sheet(workbook, holidays)
    _write_summary_sheet(workbook, project)

    logger.info("Built a plan sheet of %d row(s) and %d week(s) for %r",
                len(tasks), weeks, project.name)
    return workbook


# ---------------------------------------------------------------------------
# Public entry points
# ---------------------------------------------------------------------------

def generate_xlsx_bytes(project: Project) -> Optional[bytes]:
    """
    Generate XLSX content as bytes from a Project.

    RETURNS:
    --------
    Optional[bytes]
        The workbook, or None when it could not be built.
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl is required for Excel export")
        logger.warning("Install it with: pip install openpyxl")
        return None

    try:
        from io import BytesIO

        workbook = _create_tasks_workbook(project)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
    except Exception:
        logger.exception("Could not generate the XLSX for %r", project.name)
        return None


def export_project_to_xlsx(project: Project, filepath: str) -> bool:
    """
    Export a Project to Excel XLSX format.

    PARAMETERS:
    -----------
    project : Project
        The project to export.
    filepath : str
        Where to write the workbook. Parent directories are created.

    RETURNS:
    --------
    bool
        True when the file was written.

    EXAMPLE:
    --------
    >>> from gantt_app.models import Project, Task
    >>> from gantt_app.utils.xlsx_exporter import export_project_to_xlsx
    >>> from datetime import datetime, timedelta
    >>>
    >>> project = Project(name="My Project")
    >>> start = datetime(2024, 1, 1)
    >>> project.add_task(Task.create_task("Task 1", start, start + timedelta(days=5)))
    >>> export_project_to_xlsx(project, "/path/to/output.xlsx")
    True
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl is required for Excel export")
        logger.warning("Install it with: pip install openpyxl")
        return False

    temp_path = Path(f"{filepath}.tmp")
    try:
        path = Path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = _create_tasks_workbook(project)
        workbook.save(temp_path)
        if path.exists():
            os.replace(path, Path(f"{filepath}.bak"))
        os.replace(temp_path, path)
        logger.info("Exported %r to %s", project.name, filepath)
        return True
    except Exception:
        logger.exception("Could not export %r to %s", project.name, filepath)
        return False
    finally:
        try:
            temp_path.unlink(missing_ok=True)
        except Exception:
            pass
